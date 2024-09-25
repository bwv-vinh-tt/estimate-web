import os
from io import BytesIO
from typing import Tuple

# from sklearn.preprocessing import StandardScaler , MinMaxScaler
import cufflinks as cf
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import plotly.io as pio
import pydash as py_
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from sklearn import metrics
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.linear_model import LinearRegression
from sklearn.model_selection import train_test_split

import joblib
from apps.service.csvProcess import initCSV

from .common.constant import (
    CONST_LABEL_ASSIGNEE,
    CONST_LABEL_CODING,
    CONST_LABEL_ESTIMATED_TIME,
    CONST_LABEL_EXPECTED,
    CONST_LABEL_ISSUE_NUMBER,
    CONST_LABEL_MOD,
    CONST_LABEL_NEW,
    CONST_LABEL_NEW_MOD_FROM_INPUT,
    CONST_LABEL_TRACKER_FROM_INPUT,
    CONST_LABEL_TRAIN,
    CONST_LABEL_TRANSLATION,
    FILE_NAME,
    FILE_NAME_JOB_LIB,
    MIN_ESTIMATE_TIME,
)
from .utils import convert_data_redmine, getUrlRedmine

labelTrain = CONST_LABEL_TRAIN


def estimate(data):
    transfer = transferDataFromRequest(data)
    return loadModel(
        transfer,
        transfer[CONST_LABEL_TRACKER_FROM_INPUT],
        transfer[CONST_LABEL_NEW_MOD_FROM_INPUT],
    )


def estimateCSVProcess(file):
    df = initCSV(file, False)
    DF_CODING_NEW = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_CODING)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_NEW)
    ]
    DF_CODING_MOD = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_CODING)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_MOD)
    ]
    DF_TRANSLATION_NEW = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_TRANSLATION)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_NEW)
    ]
    DF_TRANSLATION_MOD = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_TRANSLATION)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_MOD)
    ]
    predictions = []

    # add one value '#' to LabelTrain
    labelTrain_csv = labelTrain.copy()
    labelTrain_csv.append(CONST_LABEL_ISSUE_NUMBER)
    # Check DF is empty or not
    if not DF_CODING_NEW.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_CODING, CONST_LABEL_NEW))
        )
        for index, row in DF_CODING_NEW[labelTrain_csv].iterrows():
            obj = {
                CONST_LABEL_ISSUE_NUMBER: str(row[CONST_LABEL_ISSUE_NUMBER]).split(".")[
                    0
                ],
                "prediction": "",
            }
            row.drop(CONST_LABEL_ISSUE_NUMBER, inplace=True)
            prediction = lm.predict(row.values.reshape(1, -1))
            obj["prediction"] = (
                prediction[0][0] / 60
                if prediction[0][0] / 60 > 0
                else MIN_ESTIMATE_TIME
            )
            predictions.append(obj)

    if not DF_CODING_MOD.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_CODING, CONST_LABEL_MOD))
        )
        for index, row in DF_CODING_MOD[labelTrain_csv].iterrows():
            obj = {
                CONST_LABEL_ISSUE_NUMBER: str(row[CONST_LABEL_ISSUE_NUMBER]).split(".")[
                    0
                ],
                "prediction": "",
            }
            row.drop(CONST_LABEL_ISSUE_NUMBER, inplace=True)
            prediction = lm.predict(row.values.reshape(1, -1))
            obj["prediction"] = (
                prediction[0][0] / 60
                if prediction[0][0] / 60 > 0
                else MIN_ESTIMATE_TIME
            )
            predictions.append(obj)

    if not DF_TRANSLATION_NEW.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_TRANSLATION, CONST_LABEL_NEW))
        )
        for index, row in DF_TRANSLATION_NEW[labelTrain_csv].iterrows():
            obj = {
                CONST_LABEL_ISSUE_NUMBER: str(row[CONST_LABEL_ISSUE_NUMBER]).split(".")[
                    0
                ],
                "prediction": "",
            }
            row.drop(CONST_LABEL_ISSUE_NUMBER, inplace=True)
            prediction = lm.predict(row.values.reshape(1, -1))
            obj["prediction"] = (
                prediction[0][0] / 60
                if prediction[0][0] / 60 > 0
                else MIN_ESTIMATE_TIME
            )
            predictions.append(obj)

    if not DF_TRANSLATION_MOD.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_TRANSLATION, CONST_LABEL_MOD))
        )
        for index, row in DF_TRANSLATION_MOD[labelTrain_csv].iterrows():
            obj = {
                CONST_LABEL_ISSUE_NUMBER: str(row[CONST_LABEL_ISSUE_NUMBER]).split(".")[
                    0
                ],
                "prediction": "",
            }
            row.drop(CONST_LABEL_ISSUE_NUMBER, inplace=True)
            prediction = lm.predict(row.values.reshape(1, -1))
            obj["prediction"] = (
                prediction[0][0] / 60
                if prediction[0][0] / 60 > 0
                else MIN_ESTIMATE_TIME
            )
            predictions.append(obj)

    return predictions


def loadModel(transfer, tracker, new_mod):
    Test = pd.DataFrame([transfer])

    lm = joblib.load(FILE_NAME_JOB_LIB(FILE_NAME(tracker, new_mod)))
    # Scaler
    # scaler = StandardScaler()
    # X_test_scaled = scaler.fit_transform(Test[labelTrain])
    predictions = lm.predict(Test[labelTrain])

    return predictions[0]


def trainModel():
    """
    Train model base on TRACKER and NEW_MOD
    Divide 4 files: coding_new_joblib, coding_mod_joblib, translation_new_joblib, translation_mod_joblib
    """
    ESTIMATE_UTC = pd.read_csv(os.getcwd() + "/resources/data.csv")
    ESTIMATE_UTC = ESTIMATE_UTC.drop_duplicates()
    result_effective = []

    for tracker in [CONST_LABEL_CODING, CONST_LABEL_TRANSLATION]:
        for new_mod in [CONST_LABEL_NEW, CONST_LABEL_MOD]:
            df = ESTIMATE_UTC[
                (ESTIMATE_UTC[CONST_LABEL_TRACKER_FROM_INPUT] == tracker)
                & (ESTIMATE_UTC[CONST_LABEL_NEW_MOD_FROM_INPUT] == new_mod)
            ]
            if check_df_can_train(df) == False:
                continue
            X = df[labelTrain]
            y = df[[CONST_LABEL_EXPECTED]]

            rs1, rs2 = readCSVAndTrainModel(X, y, tracker, new_mod)
            result_effective.append({"coeff": rs1, "json": rs2})
    return result_effective


def check_df_can_train(df: pd.DataFrame):
    """
    Check df can train or not
    """
    min_required = 3
    if df is None or df.empty:
        return False
    if len(df.count()[df.count() >= min_required].index) == len(df.columns):
        return True
    else:
        return False


def readCSVAndTrainModel(X: pd.DataFrame, y: pd.DataFrame, tracker: str, new_mod: str):
    """
    Train by lm or Lasso
    """
    result_effective = ""
    dict_result = {}
    try:
        X_train, X_test, y_train, y_test = train_test_split(
            X, y, test_size=0.3, random_state=10
        )

        lm = LinearRegression()
        lm.fit(X_train, y_train)

        # lasso_model = Lasso(alpha=0.1)
        # lasso_model.fit(X_train, y_train)

        joblib.dump(lm, FILE_NAME_JOB_LIB(FILE_NAME(tracker, new_mod)))
        # StandardScaler apply  X_test
        # X_test_scaled = scaler.transform(X_test)
        # predictions = lm.predict(X_test_scaled)
        predictions = lm.predict(X_test)
        sns.histplot(y_test - predictions)
        plt.scatter(y_test, predictions)

        effective = metrics.explained_variance_score(y_test, predictions)
        # print('MAE:', metrics.mean_absolute_error(y_test, predictions))
        # print('MSE:', metrics.mean_squared_error(y_test, predictions))
        print(
            f"RMSE_{tracker}_{new_mod}:",
            np.sqrt(metrics.mean_squared_error(y_test, predictions)),
        )
        print(f"Effective_{tracker}_{new_mod}: ", effective)
        result_effective = f"Effective_{tracker}_{new_mod} = " + str(effective)

        # fields effect values
        coeff_df = pd.DataFrame(lm.coef_.T, X.columns, columns=["Coefficient"])
        # print(f'COEFF_{tracker}_{new_mod}:', coeff_df)
        dict_result = coeff_df.to_dict(orient="index")
        cf.go_offline()
        pio.renderers.default = "colab"
        df = pd.DataFrame(
            {
                "EST": np.array(predictions).flatten(),
                "Spent": np.array(y_test).flatten(),
            }
        )
        fig = df.iplot(
            kind="scatter", mode="markers", asFigure=True, color=["green", "red"]
        )

        # Save the chart as a PNG file
        directorySaveImg = os.path.join(
            os.getcwd(),
            "apps",
            "static",
            "assets",
            "images",
            f"chart_{tracker}_{new_mod}.png",
        )
        pio.write_image(fig, directorySaveImg, format="png", engine="kaleido")
    except Exception as error:
        print(error)
    return result_effective, dict_result


def transferDataFromRequest(data):
    result = {}
    for key in data.keys():
        result[key] = data[key]
    return convert_data_redmine(result)


def analyzeDataFromCSVWithSpentTime(file):
    df = initCSV(file, False)
    DF_CODING_NEW = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_CODING)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_NEW)
    ]
    DF_CODING_MOD = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_CODING)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_MOD)
    ]
    DF_TRANSLATION_NEW = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_TRANSLATION)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_NEW)
    ]
    DF_TRANSLATION_MOD = df[
        (df[CONST_LABEL_TRACKER_FROM_INPUT] == CONST_LABEL_TRANSLATION)
        & (df[CONST_LABEL_NEW_MOD_FROM_INPUT] == CONST_LABEL_MOD)
    ]
    predictions = []

    # add one value '#' to LabelTrain
    labelTrain_csv = labelTrain.copy()
    labelTrain_csv.append(CONST_LABEL_ISSUE_NUMBER)
    labelTrain_csv.append(CONST_LABEL_ASSIGNEE)
    labelTrain_csv.append(CONST_LABEL_TRACKER_FROM_INPUT)
    labelTrain_csv.append(CONST_LABEL_EXPECTED)
    labelTrain_csv.append(py_.snake_case(CONST_LABEL_ESTIMATED_TIME))

    # Check DF is empty or not
    if not DF_CODING_NEW.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_CODING, CONST_LABEL_NEW))
        )
        for index, row in DF_CODING_NEW[labelTrain_csv].iterrows():
            predictions.append(convertToDfAndPredict(row, lm))

    if not DF_CODING_MOD.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_CODING, CONST_LABEL_MOD))
        )
        for index, row in DF_CODING_MOD[labelTrain_csv].iterrows():
            predictions.append(convertToDfAndPredict(row, lm))

    if not DF_TRANSLATION_NEW.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_TRANSLATION, CONST_LABEL_NEW))
        )
        for index, row in DF_TRANSLATION_NEW[labelTrain_csv].iterrows():
            predictions.append(convertToDfAndPredict(row, lm))

    if not DF_TRANSLATION_MOD.empty:
        lm = joblib.load(
            FILE_NAME_JOB_LIB(FILE_NAME(CONST_LABEL_TRANSLATION, CONST_LABEL_MOD))
        )
        for index, row in DF_TRANSLATION_MOD[labelTrain_csv].iterrows():
            predictions.append(convertToDfAndPredict(row, lm))

    return predictions


def getObjFromRowData(row):
    return {
        CONST_LABEL_ISSUE_NUMBER: str(row[CONST_LABEL_ISSUE_NUMBER]).split(".")[0],
        "prediction": "",
        CONST_LABEL_ASSIGNEE: row[CONST_LABEL_ASSIGNEE],
        CONST_LABEL_TRACKER_FROM_INPUT: row[CONST_LABEL_TRACKER_FROM_INPUT],
        CONST_LABEL_EXPECTED: row[CONST_LABEL_EXPECTED] / 60,
        "user_estimate_time": row[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)],
    }


def convertToDfAndPredict(row: pd.Series, lm: LinearRegression):
    obj = getObjFromRowData(row)
    features = row[labelTrain]
    dataFramePredict = pd.DataFrame([features])
    dataFramePredict.columns = labelTrain
    prediction = lm.predict(dataFramePredict)
    obj["prediction"] = (
        prediction[0][0] / 60 if prediction[0][0] / 60 > 0 else MIN_ESTIMATE_TIME
    )
    obj["prediction"] = round(obj["prediction"], 2)
    obj["gap"] = (
        round(
            (
                abs(obj[CONST_LABEL_EXPECTED] - obj["prediction"])
                / obj[CONST_LABEL_EXPECTED]
            ),
            2,
        )
        if obj[CONST_LABEL_EXPECTED] != 0
        else ""
    )
    obj[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)] = obj["user_estimate_time"]
    obj["gapEstimate"] = (
        round(
            (
                abs(obj[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)] - obj["prediction"])
                / obj[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)]
            ),
            2,
        )
        if obj[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)] != 0
        else ""
    )
    return obj


def exportExcelReportGap(data: list):
    wb = load_workbook(
        filename=os.path.join(os.getcwd(), "resources", "template-report-gap.xlsx")
    )
    ws = wb.active
    start_row = 2
    number_style = NamedStyle(name="number", number_format="0.00")

    for row_data in data:
        # #Column issue ID
        cell1 = ws.cell(
            row=start_row, column=1, value=row_data[CONST_LABEL_ISSUE_NUMBER]
        )
        cell1.hyperlink = getUrlRedmine(row_data[CONST_LABEL_ISSUE_NUMBER])
        cell1.style = "Hyperlink"

        # Standard Estimation
        ws.cell(row=start_row, column=2, value=row_data["prediction"]).style = (
            number_style
        )

        # Estimated time
        ws.cell(
            row=start_row,
            column=3,
            value=row_data[py_.snake_case(CONST_LABEL_ESTIMATED_TIME)],
        ).style = number_style

        # Gap * 100% (EST)
        ws.cell(
            row=start_row,
            column=4,
            value=f'=IFERROR((ABS(B{start_row}-C{start_row})/C{start_row}), "")',
        ).style = number_style
        # Gap between Standard Estimation & Estimated time (hours)
        ws.cell(
            row=start_row, column=5, value=f"=ABS(B{start_row} - C{start_row})"
        ).style = number_style

        # Spent time
        ws.cell(row=start_row, column=6, value=row_data[CONST_LABEL_EXPECTED]).style = (
            number_style
        )

        # Gap * 100% (ST)
        ws.cell(
            row=start_row,
            column=7,
            value=f'=IFERROR((ABS(B{start_row}-F{start_row})/F{start_row}), "")',
        ).style = number_style

        # Gap between Standard Estimation & Spent time (hours)
        ws.cell(
            row=start_row, column=8, value=f"=ABS(B{start_row} - F{start_row})"
        ).style = number_style

        # Tracker
        ws.cell(row=start_row, column=9, value=row_data[CONST_LABEL_TRACKER_FROM_INPUT])

        # Assignee
        ws.cell(row=start_row, column=10, value=row_data[CONST_LABEL_ASSIGNEE])

        start_row += 1  # next row

    output = BytesIO()
    wb.save(output)
    # Return value IOByte
    return output
